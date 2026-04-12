from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class MeterSheetDefinition:
    worksheet_name: str
    location_key: str
    rows: tuple[int, int]
    date_column: str
    streams: tuple[tuple[str, str, str], ...]


@dataclass
class ParsedReading:
    worksheet_name: str
    location_key: str
    category_name: str
    meter_label: str
    reading_date: date
    value: float
    notes: str


SHEET_DEFINITIONS: tuple[MeterSheetDefinition, ...] = (
    MeterSheetDefinition(
        worksheet_name="AP 12",
        location_key="ap12",
        rows=(4, 1000),
        date_column="F",
        streams=(
            ("Energy", "Main Electricity Meter", "G"),
        ),
    ),
    MeterSheetDefinition(
        worksheet_name="AP 12",
        location_key="ap12",
        rows=(4, 1000),
        date_column="Q",
        streams=(
            ("Cold Water", "Kitchen Cold Water", "R"),
            ("Hot Water", "Kitchen Hot Water", "T"),
            ("Cold Water", "Bathroom Cold Water", "V"),
            ("Hot Water", "Bathroom Hot Water", "X"),
        ),
    ),
    MeterSheetDefinition(
        worksheet_name="AP 12",
        location_key="ap12",
        rows=(4, 1000),
        date_column="AD",
        streams=(
            ("Heating", "Kitchen Radiator", "AE"),
            ("Heating", "Living Room Radiator", "AG"),
            ("Heating", "Big Bedroom Radiator", "AI"),
            ("Heating", "Small Bedroom Radiator", "AK"),
        ),
    ),
    MeterSheetDefinition(
        worksheet_name="AP 15",
        location_key="ap15",
        rows=(4, 1000),
        date_column="F",
        streams=(
            ("Energy", "Main Electricity Meter", "G"),
        ),
    ),
    MeterSheetDefinition(
        worksheet_name="AP 15",
        location_key="ap15",
        rows=(4, 1000),
        date_column="Q",
        streams=(
            ("Gas", "Main Gas Meter", "R"),
        ),
    ),
    MeterSheetDefinition(
        worksheet_name="AP 15",
        location_key="ap15",
        rows=(4, 1000),
        date_column="Z",
        streams=(
            ("Cold Water", "Kitchen Cold Water", "AA"),
            ("Hot Water", "Kitchen Hot Water", "AC"),
            ("Cold Water", "Bathroom Cold Water", "AE"),
            ("Hot Water", "Bathroom Hot Water", "AG"),
        ),
    ),
)


def _normalize_location_key(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


def normalize_location_name(value: str) -> str:
    return _normalize_location_key(value)


def _coerce_date(value) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned == "-":
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _coerce_number(value) -> Optional[float]:
    if value in (None, "", "-"):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned == "-":
            return None
        cleaned = cleaned.replace(" ", "")
        try:
            return float(Decimal(cleaned.replace(",", ".")))
        except (InvalidOperation, ValueError):
            return None
    return None


def _iter_readings(ws: Worksheet, definition: MeterSheetDefinition) -> Iterable[ParsedReading]:
    row_start, row_end = definition.rows
    for row in range(row_start, row_end + 1):
        reading_date = _coerce_date(ws[f"{definition.date_column}{row}"].value)
        if reading_date is None:
            continue

        for category_name, meter_label, value_column in definition.streams:
            value = _coerce_number(ws[f"{value_column}{row}"].value)
            if value is None:
                continue
            yield ParsedReading(
                worksheet_name=definition.worksheet_name,
                location_key=definition.location_key,
                category_name=category_name,
                meter_label=meter_label,
                reading_date=reading_date,
                value=value,
                notes=f"Imported from {definition.worksheet_name} workbook history",
            )


def parse_meter_workbook(file_bytes: bytes) -> List[ParsedReading]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    parsed: List[ParsedReading] = []
    seen_keys: set[tuple[str, str, str, date]] = set()

    for definition in SHEET_DEFINITIONS:
        if definition.worksheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[definition.worksheet_name]
        for reading in _iter_readings(worksheet, definition):
            dedupe_key = (
                reading.location_key,
                reading.category_name.lower(),
                reading.meter_label.lower(),
                reading.reading_date,
            )
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            parsed.append(reading)

    parsed.sort(key=lambda item: (item.location_key, item.category_name, item.meter_label, item.reading_date))
    return parsed
